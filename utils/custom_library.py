from robot.api.deco import keyword
import os
import string
import pandas as pd
import imaplib
import email
from email.header import decode_header
import re
import random
from datetime import datetime
from configparser import ConfigParser
from email.utils import parseaddr

import mysql.connector
from mysql.connector import errorcode
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.utils.exceptions import InvalidFileException
import pyautogui
from dotenv import load_dotenv


class TestRunManager:
    """
    A utility class for managing test execution metadata, user data storage,
    database configuration loading from both config.ini and .env files.
    """
    def __init__(self):
        """
        Initialize configuration variables from .env and config.ini, then set up DB connection details.
        """
        try:
            # Step 1: Construct the full path to the config.ini file
            self.config_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), '..', 'configs', 'config.ini')

            # Step 2: Load .env file (if present)
            env_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), '..', '.env')
            if os.path.exists(env_path):
                load_dotenv(dotenv_path=env_path)

            # Step 3: Read and override from config.ini if available
            self.database_config = {}
            self.read_config()

            # Step 4: Initialize DB connection and cursor to None
            self.connection = None
            self.cursor = None

            # Step 5: Extract DB config values into instance variables
            self.database_username = self.database_config["username"]
            self.database_password = self.database_config["password"]
            self.database_host = self.database_config["host"]
            self.database_port = self.database_config["port"]
            self.database_name = self.database_config["database"]
            self.execution_status_start = self.database_config["execution_status_start"]
            self.execution_status_end = self.database_config["execution_status_end"]

            # Step 6: Default DB values from .env (override by config.ini if found)
            self.register_email_app_password = os.getenv("REGISTER_EMAIL_APP_PASSWORD")
            self.imap_server = os.getenv("IMAP_SERVER")
            self.email_subject = os.getenv("SUBJECT")

        except Exception as e:
            print(f"[ERROR] Initialization failed: {e}")

    def read_config(self):
        """
        Reads configuration from the config.ini file.
        """
        # Step 1: Check if config file exists
        if not os.path.exists(self.config_file):
            print(f"Config file not found: {self.config_file}")
            return

        # Step 2: Create ConfigParser instance and read the file
        config = ConfigParser()
        try:
            with open(self.config_file, 'r', encoding='utf-8') as file:
                config.read_file(file)  # Read file with correct encoding

            # Step 3: Validate section presence
            if "mysql" not in config:
                print("Error: 'mysql' section not found in config.ini")
                return

            # Step 4: Populate DB config dictionary with fallback defaults
            self.database_config = {
                "username": config.get("mysql", "username", fallback=None),
                "password": config.get("mysql", "password", fallback=None),
                "host": config.get("mysql", "host", fallback=None),
                "port": config.get("mysql", "port", fallback="3306"),
                "database": config.get("mysql", "database", fallback=None),
                "execution_status_start": config.get("mysql", "execution_status_start", fallback="STARTED"),
                "execution_status_end": config.get("mysql", "execution_status_end", fallback="COMPLETED"),
            }

            if not all(self.database_config.values()):
                print("Error: Missing required database configuration values.")

        except Exception as e:
            print(f"[ERROR] Failed to read configuration: {e}")

    def connect_db(self):
        """
        Establish connection to the database.
        - Creates the database if it doesn't exist.
        - Ensures required tables are created.
        """
        # Step 1: Check if essential DB config values exist
        if not all([self.database_username, self.database_password, self.database_host, self.database_name]):
            print("Database credentials are missing. Check your config.ini file.")
            return False

        try:
            # Step 2: Connect to MySQL server without selecting DB
            temp_connection = mysql.connector.connect(
                user=self.database_username,
                password=self.database_password,
                host=self.database_host,
                port=int(self.database_port),
                database=self.database_name,
                auth_plugin="mysql_native_password"
            )

            temp_cursor = temp_connection.cursor()

            # Step 3: Create DB if not already present
            temp_cursor.execute(f"CREATE DATABASE IF NOT EXISTS {self.database_name}")
            temp_cursor.close()
            temp_connection.close()

            # Step 4: Now connect to the actual target database
            self.connection = mysql.connector.connect(
                user=self.database_username,
                password=self.database_password,
                host=self.database_host,
                port=int(self.database_port),
                database=self.database_name,
                auth_plugin="mysql_native_password"  # Ensures compatibility with modern MySQL versions
            )

            self.cursor = self.connection.cursor()
            print("Database connection successful.")

            # Step 5: Ensure required tables exist before proceeding
            self.ensure_tables_exist()
            return True

        except mysql.connector.Error as err:
            print(f"[ERROR] Database connection failed: {err}")
            return False

    def ensure_tables_exist(self):
        """
        Create essential tables if they do not exist.
        - test_execution_reports
        - user_data
        """
        # Step 1: Define table creation SQL for both tables
        tables = {
            "test_execution_reports": """
                CREATE TABLE IF NOT EXISTS test_execution_reports (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255),
                    execution_start_time DATETIME,
                    execution_end_time DATETIME NULL,
                    execution_status VARCHAR(50)
                )
            """,
            "users": """
                CREATE TABLE IF NOT EXISTS user_data (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    business_name VARCHAR(255),
                    username VARCHAR(255),
                    password VARCHAR(255),
                    email VARCHAR(255)
                )
            """
        }

        # Step 2: Execute table creation if not found
        try:
            for table_name, create_query in tables.items():
                self.cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
                result = self.cursor.fetchone()

                if not result:
                    self.cursor.execute(create_query)
                    self.connection.commit()
                    print(f"Table '{table_name}' created successfully.")
                else:
                    print(f"Table '{table_name}' already exists.")


        except Exception as e:
            print(f"[ERROR] Failed to ensure tables exist: {e}")
            raise

    def insert_start_time(self):
        """
        Insert test Execution Start Time into database.
        """
        # Step 1: Ensure DB is connected
        if not self.connect_db():
            print("Skipping database insertion due to connection failure.")
            return

        # Step 2: Get current datetime
        current_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Step 3: Insert a new test execution row
        try:
            sql = "INSERT INTO test_execution_reports (Name, execution_start_time, execution_status) VALUES (%s, %s, %s)"
            self.cursor.execute(sql, ('OpenCart', current_timestamp, self.execution_status_start))
            self.connection.commit()
            print("Start time inserted successfully.")

        except Exception as e:
            print(f"[ERROR] Failed to insert start time: {e}")
            raise

    def update_end_time(self):
        """
        Update the last test execution with an end time and completion status.
        """
        # Step 1: Ensure DB is connected
        if not self.connect_db():
            print("Skipping database update due to connection failure.")
            return

        # Step 2: Get current datetime
        current_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            # Step 3: Get last inserted row
            self.cursor.execute("SELECT id FROM test_execution_reports ORDER BY id DESC LIMIT 1")
            last_id = self.cursor.fetchone()

            if not last_id:
                print("No previous test reports found to update.")
                return

            # Step 4: Update end time and status
            sql = "UPDATE test_execution_reports SET execution_end_time = %s, execution_status = %s WHERE id = %s"
            self.cursor.execute(sql, (current_timestamp, self.execution_status_end, last_id[0]))
            self.connection.commit()
            print("End time updated successfully.")

        except Exception as e:
            print(f"[ERROR] Failed to update end time: {e}")
            raise

    @staticmethod
    def save_user_data_to_excel(firstname, lastname, email, telephone, password, confirm_password, excel_path):
        """
        Save user registration data into an Excel file.
        Appends to existing file; does not overwrite previous entries.
        """
        try:
            # Step 1: Define the directory where the Excel file will be stored
            directory = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'utils'))
            os.makedirs(directory, exist_ok=True)  # Step 2: Create the directory if it doesn't exist

            # Step 3: Create full path to Excel file
            full_path = os.path.join(directory, os.path.basename(excel_path))

            # Step 4: If Excel file exists, load it; otherwise, create a new one with headers
            if os.path.exists(full_path):
                workbook = openpyxl.load_workbook(full_path)
                sheet = workbook.active
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                # Add headers only once when creating a new file
                sheet.append(
                    ['First Name', 'Last Name', 'Email', 'Telephone', 'Password', 'Confirm Password', 'Created At'])

            # Step 5: Append the user data to the next available row
            sheet.append([
                firstname,
                lastname,
                email,
                telephone,
                password,
                confirm_password,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ])

            # Step 6: Save the updated workbook
            workbook.save(full_path)
            print(f"User credentials saved to Excel: {full_path}")

            return {"status": "error", "message": "User data saved successfully to Excel.", "file_path": None}

        except Exception as e:
            error_message = f"[ERROR] Failed to save user data to Excel: {e}"
            print(error_message)
            return {"status": "error", "message": str(e), "file_path": None}

    def save_user_data_to_db(self, firstname, lastname, email, telephone, password, confirm_password):
        """
        Save user registration data into the MySQL database.
        Ensures database and table creation before insertion.
        Returns a dictionary with success status and message.
        """
        try:
            # Step 1: Connect without specifying database to create DB if needed
            temp_connection = mysql.connector.connect(
                user=self.database_username,
                password=self.database_password,
                host=self.database_host,
                port=int(self.database_port),
                auth_plugin="mysql_native_password"
            )
            temp_cursor = temp_connection.cursor()

            # Step 2: Create the database if it does not exist
            temp_cursor.execute(f"CREATE DATABASE IF NOT EXISTS {self.database_name}")
            temp_cursor.close()
            temp_connection.close()

            # Step 3: Now connect to the actual target database
            self.connection = mysql.connector.connect(
                user=self.database_username,
                password=self.database_password,
                host=self.database_host,
                port=int(self.database_port),
                database=self.database_name,
                auth_plugin="mysql_native_password"
            )
            self.cursor = self.connection.cursor()
            print("Database connection successful.")

            # Step 4: Create table if not exists
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS registered_users (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    firstname VARCHAR(100) NOT NULL,
                    lastname VARCHAR(100) NOT NULL,
                    email VARCHAR(255) NOT NULL UNIQUE,
                    telephone VARCHAR(15),
                    password TEXT NOT NULL,
                    confirm_password TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Step 5: Insert user data
            self.cursor.execute("""
                INSERT INTO registered_users (firstname, lastname, email, telephone, password, confirm_password)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (firstname, lastname, email, telephone, password, confirm_password))

            self.connection.commit()
            message = f"User credentials saved to DB: {firstname} {lastname}"
            print(message)
            return {"success": True, "message": message}

        except mysql.connector.IntegrityError as ie:
            message = f"[ERROR] Duplicate email entry: {email} - {ie}"
            print(message)
            return {"success": False, "message": message}

        except Exception as e:
            message = f"[ERROR] Failed to save user data to database: {e}"
            print(message)
            return {"success": False, "message": message}

        finally:
            # Step 6: Only close if the connection was successfully created
            if hasattr(self, "connection") and self.connection and self.connection.is_connected():
                self.cursor.close()
                self.connection.close()

    @keyword("Fetch Registration Login Link")
    def get_login_link_from_email(self, email_user, email_pass):
        """
        Searches accessible email folders for a message with the expected subject.
        Extracts login link and sender email.

        Args:
            email_user (str): Gmail address.
            email_pass (str): Gmail App Password (not your actual password).

        Returns:
            tuple: (login_link, sender_email) if found, otherwise (None, None)
        """
        mail = None
        try:
            # Step 1: Connect and log in to Gmail IMAP
            mail = imaplib.IMAP4_SSL(self.imap_server)

            # Strip any invisible characters
            email_user = email_user.strip()
            email_pass = email_pass.strip()

            mail.login(email_user, email_pass)
            print("Login successful")

            # Step 2: Get available folders
            status, folders = mail.list()
            if status != "OK":
                print("[ERROR] Unable to list folders.")
                return None, None

            for folder_bytes in folders:
                # Extract folder name from response
                folder = folder_bytes.decode().split(' "/" ')[-1].strip('"')
                print(f"Checking folder: {folder}")

                # Step 3: Select folder safely
                status, _ = mail.select(f'"{folder}"')
                if status != "OK":
                    print(f"[ERROR] Cannot select folder: {folder}")
                    continue

                # Step 4: Search for emails with the given subject
                result, data = mail.search(None, 'ALL')
                if result != "OK":
                    print(f"[ERROR] No matching emails found in {folder}")
                    continue

                email_ids = data[0].split()[-20:]

                for eid in reversed(email_ids):  # Check latest first
                    # Fetch full email
                    result, msg_data = mail.fetch(eid, "(RFC822)")
                    if result != "OK":
                        continue

                    raw_email = msg_data[0][1]
                    msg = email.message_from_bytes(raw_email)

                    # Decode subject
                    raw_subject = msg.get("Subject", "")
                    decoded_subject, enc = decode_header(raw_subject)[0]
                    if isinstance(decoded_subject, bytes):
                        decoded_subject = decoded_subject.decode(enc or "utf-8", errors="ignore")

                    print(f"Found subject: {decoded_subject}")
                    if "thank you for registering" not in decoded_subject.lower():
                        continue

                    from_email = msg.get("From")
                    sender_email = email.utils.parseaddr(from_email)
                    print(f"Sender email found: {sender_email}")

                    # Extract body
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() in ["text/plain", "text/html"] and "attachment" not in str(
                                    part.get("Content-Disposition")):
                                charset = part.get_content_charset() or "utf-8"
                                body = part.get_payload(decode=True).decode(charset, errors="ignore")
                                break
                    else:
                        charset = msg.get_content_charset() or "utf-8"
                        body = msg.get_payload(decode=True).decode(charset, errors="ignore")

                    print("Body preview:\n", body[:300])  # Optional preview

                    # Find login link
                    match = re.search(r"https?://\S+?route=account/login", body)
                    if match:
                        login_link = match.group(0)
                        print(f"Login link found: {login_link}")
                        return login_link, sender_email

            print("No login link found in any folder.")
            return None, None

        except Exception as e:
            print(f"[ERROR] Email check failed: {e}")
            return None, None

        finally:
            if mail:
                mail.logout()

    def store_user_data(self, username="tester", password="&lackMan123!", business_name="Demo", email_addr="demo@gmail.com"):
        """
        Store a user's credentials into the 'user_data' table.
        """
        # Step 1: Ensure DB is connected
        if not self.connect_db():
            print("Skipping database insertion due to connection failure.")
            return

        try:
            ## Hash the password before storing
            # hashed_password = bcrypt.hashpw(password.encode('utf-8'),bcrypt.gensalt()).decode('utf-8')

            # Step 2: Insert user data
            sql = "INSERT INTO user_data (business_name, username, password, email) VALUES (%s, %s, %s, %s)"
            self.cursor.execute(sql, (business_name, username, password, email_addr))
            self.connection.commit()
            print("User data saved successfully to database with encrypted password.")

        except Exception as e:
            print(f"[ERROR] Failed to store user data: {e}")

    # def create_user_excel(self, firstname=None, lastname=None, email=None, telephone=None, password=None, confirm_password=None):
    #     try:
    #         wb = Workbook()
    #         ws = wb.active
    #         ws.title = "UserData"
    #
    #         # Set headers
    #         headers = ["Firstname", "Lastname", "Email", "Telephone", "Password", "Confirm Password"]
    #         ws.append(headers)
    #
    #         # Add user data with default None
    #         user_data = [firstname, lastname, email, telephone, password, confirm_password]
    #         ws.append(user_data)
    #
    #         # Save the workbook
    #         wb.save(self.filename)
    #         print(f"Excel file '{self.filename}' created successfully.")
    #
    #     except InvalidFileException:
    #         print(f"Error: Invalid Excel file format. Please ensure the file '{self.filename}' is not corrupted.")
    #     except PermissionError:
    #         print(
    #             f"Error: Permission denied when trying to save '{self.filename}'. Make sure the file is not open elsewhere.")
    #     except Exception as e:
    #         print(f"An unexpected error occurred: {e}")

    @staticmethod
    def generate_random_string(length=10, special_chars=False):
        """
        Generate a random string with a specified length and optional special characters.
        """
        try:
            # Step 1: Define allowed characters
            chars = string.ascii_letters + string.digits + ("!@#$&*?" if special_chars else "")

            while True:
                # Step 2: Create a random string
                random_string = ''.join(random.choices(chars, k=length))

                # Step 3: Ensure at least one special character if required
                if not special_chars or any(c in "!@#$&*?" for c in random_string):
                    return random_string
        except Exception as e:
            print(f"[ERROR] Failed to generate random string: {e}")
            return None

    @staticmethod
    def generate_unique_username(base_name, length=6):
        """
        Generate a unique username using a base name and a random numeric suffix.
        """
        # Step 1: Create username by appending random digits
        try:
            random_suffix = ''.join(random.choices(string.digits, k=length))
            return f"{base_name}_{random_suffix}"
        except Exception as e:
            print(f"[ERROR] Failed to generate username: {e}")
            return None

    @staticmethod
    def generate_random_email():
        """Generate a random email address."""
        # Step 1: Generate 10-character random string + domain
        try:
            return f"{''.join(random.choices(string.ascii_lowercase + string.digits, k=10))}@example.com"
        except Exception as e:
            print(f"[ERROR] Failed to generate email: {e}")
            return None

    @staticmethod
    def zoom_in():
        """Zoom in using Ctrl + '+' shortcut."""
        try:
            for _ in range(2):  # Step 1: Simulate zoom in 2 times
                pyautogui.hotkey('ctrl', '+')
        except Exception as e:
            print(f"[ERROR] Failed to zoom in: {e}")

    @staticmethod
    def zoom_out():
        """Zoom out using Ctrl + '-' shortcut."""
        try:
            for _ in range(2):  # Step 1: Simulate zoom out 2 times
                pyautogui.hotkey('ctrl', '-')
        except Exception as e:
            print(f"[ERROR] Failed to zoom out: {e}")

    def get_login_link(self, email_addr, password):
        mail = None
        try:
            print("Running....")
            mail = imaplib.IMAP4_SSL(self.imap_server)
            mail.login(email_addr, password)
            mail.select("inbox")

            result, data = mail.search(None, '(FROM "info@opencart.com.gr")')
            email_ids = data[0].split()
            if not email_ids:
                return None

            emails_with_dates = []
            for eid in email_ids:
                result, msg_data = mail.fetch(eid, '(BODY[HEADER.FIELDS (DATE)])')
                raw_date = msg_data[0][1].decode().strip().replace('Date: ', '')
                email_date = email.utils.parsedate_to_datetime(raw_date)
                emails_with_dates.append((eid, email_date))

            sorted_emails = sorted(emails_with_dates, key=lambda x: x[1], reverse=True)

            for email_id, _ in sorted_emails:
                result, data = mail.fetch(email_id, '(RFC822)')
                raw_email = data[0][1]
                msg = email.message_from_bytes(raw_email)

                raw_subject = msg.get("Subject", "")
                decoded_subject, enc = decode_header(raw_subject)[0]
                if isinstance(decoded_subject, bytes):
                    decoded_subject = decoded_subject.decode(enc or "utf-8", errors="ignore")

                print(f"Found subject: {decoded_subject}")
                if "thank you for registering" not in decoded_subject.lower():
                    continue

                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() in ["text/plain", "text/html"] and "attachment" not in str(
                                part.get("Content-Disposition")):
                            charset = part.get_content_charset() or "utf-8"
                            body = part.get_payload(decode=True).decode(charset, errors="ignore")
                            break
                else:
                    charset = msg.get_content_charset() or "utf-8"
                    body = msg.get_payload(decode=True).decode(charset, errors="ignore")

                print("Body preview:\n", body[:300])  # Optional preview

                match = re.search(r'https?://\S+?route=account/login', body)
                if match:
                    login_link = match.group(0)
                    print(f"Latest valid Login Link: {login_link}")
                    return login_link

            return None

        except Exception as e:
            print(f"Error fetching email: {e}")
            return None
        finally:
            if mail:
                mail.logout()


# Initialize and test the class
if __name__ == "__main__":
    test_manager = TestRunManager()
    # test_manager.insert_start_time()
    # test_manager.update_end_time()
    # test_manager.store_user_data()
    test_manager.get_login_link_from_email("vimalmuruges+F9@imetanic.co", "pjnfclygrlgsjsfj")
    # test_manager.get_login_link("vimalmuruges+F9@imetanic.co", "pjnfclygrlgsjsfj")
    # test_manager.zoom_in()
    # test_manager.zoom_out()
    # test_manager.generate_random_string()
    # test_manager.generate_random_email()
    # test_manager.generate_random_string_with_special_chars()
    # test_manager.generate_random_string_without_special_chars()
    # test_manager.generate_unique_username()
    # test_manager.save_user_data_to_excel()
    # test_manager.store_user_data()
